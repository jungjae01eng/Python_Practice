# @author Jungjae Lee
# Independent Study
# Created on May 23, 2022
# Sources: Python Tutorial by TechWorld with Nana on the Youtube channel
# --------------------------------------------------------------------------------


"""
Exercise:
    a. Read spreadsheet file and automate stuff
    b. List each company with respective product count
    c. List products with inventory less than 10
    d. List each company with respective total inventory value
    e. Write to spreadsheet: Calculate and write inventory value for each product into spreadsheet
"""


import openpyxl
# When you work with Excel, make sure to import the python package to work with spreadsheets specifically before you start coding!

# Step to install the python package:
# First, go to Python Package Website (https://pypi.org/), and look up "openpyxl" on search browse.
# Then go to the most recently updated package ("openpypl 3.0.10" for me).
# Click the copy logo next to "pip install openpypx" and paste it on the terminal in your code to install the module you need.


inv_file = openpyxl.load_workbook("Project02_inventory.xlsx")
product_list = inv_file["Sheet1"]


products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}


# "range" starts from 0 in default, NOT 1. Therefore, it will make an error if we say "range(product_list.max_row)."
for product_row in range(2, product_list.max_row + 1):       # The range is row 2 to max row (30)
    product_num = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value
    inventory_price = product_list.cell(product_row, 5)

    # Calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory + price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Add value for total inventory price
    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)


inv_file.save("Project02_inventory_with_total_value.xlsx")
